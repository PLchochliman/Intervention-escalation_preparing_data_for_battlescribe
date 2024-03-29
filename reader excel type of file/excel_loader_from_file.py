from openpyxl import load_workbook

def read_excel(name_of_sheet):
    book_opened = load_workbook(name_of_sheet)

    return book_opened #xw.load(name_of_sheet)


workbook = read_excel("Already downloaded\\Pancerni.xlsx")
print(workbook.sheetnames)


units = workbook["Escalation - jednostki"]
weapons = workbook["Escalation - Broń"]
rules = workbook["Escalation - Zasady specjalne"]
formations = workbook["Escalation - Formacje do odczyt"]

units_wide = 14
weapons_wide = 7
print(units.cell(row=1, column=1).value)

